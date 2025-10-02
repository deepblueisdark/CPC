import pandas as pd
from statistics import mean 
import openpyxl
from openpyxl.workbook import Workbook

from datetime import datetime


col_names=['ano',
           'mes',
           'dia',
           'SAE01',  ## total madeira 
           'SAE02',  ## cacho
           'SAE03',  ##Madre
           'SAE04',  ## Beni
           'SAE05',   ##mamore
           'SAE06',  ##guapore
           'SAE07',   ##inc ate JRB
           'SAE08',   ##JRB jirau
           'SAE09',   ##Jirau-SAE
           'SAE10',    ##JRB-SAE
           'SAE11',   ##montJP
           'SAE12',   ## JPVILA
           'SAE13',   ## Inc_JPVILA
           'SAE14']    


header=['ANO ',
        'MES',
        'DIA',
        'CACH_ESP',
        'PUERTO SILES',
        'MONTANTE_JACI',
        'MORADA_NOVA',
        'INC_JIRAU_SAE',
        'INC_JRB',
        'BENI CONF. MADRE DIOS',
        'GUAPORE',
        'JRB_SAE',
        'JRB',
        'MADREDIOS',
        'RIO MADEIRA',
        'bacia_inc_JPVILA',
        'bacia_JPVILA']



coluna=("A","B","C","D","E","F","G","H","I","J","K","L","M","N","O","P","Q","R","S","T","U","V","W","Y","Z")   


filename = "CHUVA_OBSERVADA.xlsx"
wb=Workbook()
sheet=wb.create_sheet("54km")
sheet.title="54km"


df = pd.read_excel("chuva_diaria_CPC.xlsx", sheet_name="chuva_media",names=col_names, header=0)
print(df)
num_cols=len(header)
num_itens=len(df)
for n in range(0,num_cols):
    celula=coluna[n]+"1"
    sheet[celula]=header[n]

for dia in range(0,num_itens):
    row=dia+2
    madeira=(df.SAE03[dia]+df.SAE04[dia]+df.SAE05[dia]+df.SAE06[dia]+df.SAE07[dia]+df.SAE08[dia]+df.SAE09[dia])/7
    celula=coluna[0]+str(row)
    sheet[celula]=df.ano[dia]   
    celula=coluna[1]+str(row)
    sheet[celula]=df.mes[dia]   
    celula=coluna[2]+str(row)
    sheet[celula]=df.dia[dia]

	
    celula=coluna[3]+str(row)
    sheet[celula]=df.SAE02[dia]   ## cachoeira   
    celula=coluna[4]+str(row)
    sheet[celula]=df.SAE05[dia]    ## mamore
    celula=coluna[5]+str(row)
    sheet[celula]=df.SAE11[dia]     ## mont_jpvilla
    celula=coluna[6]+str(row)
    sheet[celula]=df.SAE08[dia]     ### JRB-JIRAU
    celula=coluna[7]+str(row)
    sheet[celula]=df.SAE09[dia]    ### jirau-SAE
    celula=coluna[8]+str(row)
    sheet[celula]=df.SAE07[dia]     ##Inc_JRB
    celula=coluna[9]+str(row)
    sheet[celula]=df.SAE04[dia]     ###Beni
    celula=coluna[10]+str(row)
    sheet[celula]=df.SAE06[dia]   ###guapore
    celula=coluna[11]+str(row)
    sheet[celula]=df.SAE10[dia]    ###jrb_sae
    celula=coluna[12]+str(row)
    sheet[celula]=df.SAE01[dia]   ###JRB
    celula=coluna[13]+str(row)
    sheet[celula]=df.SAE03[dia]    ###madredios
    celula=coluna[14]+str(row)
    sheet[celula]=madeira   ####madeira 
    celula=coluna[15]+str(row)
    sheet[celula]=df.SAE13[dia]      ### inc_JVILA     
    celula=coluna[16]+str(row)
    sheet[celula]=df.SAE12[dia]      ### inc_JVILA     
 
    row=+1
    


wb.save(filename)


filename = "CHUVA_OBSERVADA_CPC_COMPLETA.xlsx"
wb2=openpyxl.load_workbook(filename)

sheet=wb2['54km']


for dia in range(0,num_itens):
    start0=datetime(1979,1,1,0,0)
    atual=datetime(df.ano[dia], df.mes[dia], df.dia[dia], 0, 0)
    num_dias_time=atual-start0
    print(num_dias_time)
    print(type(num_dias_time))
    print(type(atual))
    print("=======",atual)
    num_dias=str(num_dias_time)
    row=1+int(num_dias[0:6],10) 
    madeira=(df.SAE03[dia]+df.SAE04[dia]+df.SAE05[dia]+df.SAE06[dia]+df.SAE07[dia]+df.SAE08[dia]+df.SAE09[dia])/7
    celula=coluna[0]+str(row)
    sheet[celula]=df.ano[dia]   
    celula=coluna[1]+str(row)
    sheet[celula]=df.mes[dia]   
    celula=coluna[2]+str(row)
    sheet[celula]=df.dia[dia]
    celula=coluna[3]+str(row)
    sheet[celula]=0
	
    celula=coluna[4]+str(row)
    sheet[celula]=df.SAE01[dia]   ## cachoeira   
    celula=coluna[5]+str(row)
    sheet[celula]=df.SAE02[dia]    ## mamore
    celula=coluna[6]+str(row)
    sheet[celula]=df.SAE03[dia]     ## mont_jpvilla
    celula=coluna[7]+str(row)
    sheet[celula]=df.SAE04[dia]     ### JRB-JIRAU
    celula=coluna[8]+str(row)
    sheet[celula]=df.SAE05[dia]    ### jirau-SAE
    celula=coluna[9]+str(row)
    sheet[celula]=df.SAE06[dia]     ##Inc_JRB
    celula=coluna[10]+str(row)
    sheet[celula]=df.SAE07[dia]     ###Beni
    celula=coluna[11]+str(row)
    sheet[celula]=df.SAE08[dia]   ###guapore
    celula=coluna[12]+str(row)
    sheet[celula]=df.SAE09[dia]    ###jrb_sae
    celula=coluna[13]+str(row)
    sheet[celula]=df.SAE10[dia]   ###JRB
    celula=coluna[14]+str(row)
    sheet[celula]=df.SAE11[dia]    ###madredios
    celula=coluna[15]+str(row)
    sheet[celula]=madeira   ####madeira 
    celula=coluna[16]+str(row)
    sheet[celula]=df.SAE12[dia]      ### inc_JVILA     
    celula=coluna[17]+str(row)
    sheet[celula]=df.SAE13[dia]      ### inc_JVILA     
 


wb2.save(filename)


